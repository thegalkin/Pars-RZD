# -*- coding: UTF-8 -*-
from bs4 import BeautifulSoup as bs
import sys, time, re
import pathlib as ph
from os import listdir,chdir
from os.path import isfile, join
import openpyxl as opx
def appendToDb():
    return

homeDirectory = ph.Path.cwd()
chdir("{}/downloads".format(ph.Path.cwd()))
onlyfiles = [f for f in listdir(ph.Path.cwd()) if isfile(join(ph.Path.cwd(), f))] # создаем список файлов в папке загрузок

i = 0
items = []
print(onlyfiles)
for fName in onlyfiles:
    
    i+=1
    #if i == 2: break
    with open(str(join(ph.Path.cwd(), fName)), "r") as site:
        try:
            soup = bs(site.read(), "lxml")
        except UnicodeDecodeError:
            print("UnicodeDecodeError")
        """for line in soup:
            item = soup.find("td", class_="availabiItm").text
            item = re.sub("^\s+|\n|\t|\s+$", '', item)
            items.append(item)"""
        """print(soup.title.string)
        for tag in soup.findAll("td", class_="availabiItm"):
            
            print("{0}: {1}: ".format(tag.name, re.sub("\n|\t", '', tag.h3.string) ))
        print("Коляска: ")
        for tag in soup.findAll("td", class_="icoLineType1"):
            
            if tag.find(class_="pass-ico ico-color2"):
                print("25%")
            elif(tag.find(class_="pass-ico ico-color1") ):
                print("12.5%")
            elif tag.find(class_="pass-ico ico-color0") :
                print("0%")
        print("Трость: ")
        for tag in soup.findAll("td", class_="icoLineType2"):
            
            if tag.find(class_="pass-ico ico-color2"):
                print("25%")
            elif(tag.find(class_="pass-ico ico-color1") ):
                print("12.5%")
            elif tag.find(class_="pass-ico ico-color0") :
                print("0%")
        print("Слух: " )
        for tag in soup.findAll("td", class_="icoLineType3"):
            
            if tag.find(class_="pass-ico ico-color2"):
                print("25%")
            elif(tag.find(class_="pass-ico ico-color1") ):
                print("12.5%")
            elif tag.find(class_="pass-ico ico-color0") :
                print("0%")
        print("Зрение: " )
        for tag in soup.findAll("td", class_="icoLineType4"):
            
            if tag.find(class_="pass-ico ico-color2"):
                print("25%")
            elif(tag.find(class_="pass-ico ico-color1") ):
                print("12.5%")
            elif tag.find(class_="pass-ico ico-color0") :
                print("0%")"""
        
        #services and summing
        summing = []
        services = []
        
        for tag in soup.findAll("td", class_="availabiItm"):
            try:
                services.append(re.sub("\n|\t", '', tag.h3.string))
            except AttributeError:
                re.sub("\n|\t", '', tag.string)
            if tag.find(class_="percentRate pR-1"):
                summing.append(tag.find(class_="percentRate pR-1").string.replace(".", ",") )
            elif(tag.find(class_="percentRate pR-2") ):
                summing.append(tag.find(class_="percentRate pR-2").string.replace(".", ",") )
            elif tag.find(class_="percentRate pR-3") :
                summing.append(tag.find(class_="percentRate pR-3").string.replace(".", ",") )
        
            
        #wheelChair
        wheelChair = []
        for tag in soup.findAll("td", class_="icoLineType1"):
            
            if tag.find(class_="pass-ico ico-color2"):
                wheelChair.append("25%")
            elif(tag.find(class_="pass-ico ico-color1") ):
                wheelChair.append("12,5%")
            elif tag.find(class_="pass-ico ico-color0") :
                wheelChair.append("0%")

        #walkStick
        walkStick = []
        for tag in soup.findAll("td", class_="icoLineType2"):
            
            if tag.find(class_="pass-ico ico-color2"):
                walkStick.append("25%")
            elif(tag.find(class_="pass-ico ico-color1") ):
                walkStick.append("12,5%")
            elif tag.find(class_="pass-ico ico-color0") :
                walkStick.append("0%")

        #hearing
        hearing = []
        for tag in soup.findAll("td", class_="icoLineType3"):
            
            if tag.find(class_="pass-ico ico-color2"):
                hearing.append("25%")
            elif(tag.find(class_="pass-ico ico-color1") ):
                hearing.append("12,5%")
            elif tag.find(class_="pass-ico ico-color0") :
                hearing.append("0%")

        #seeing
        seeing = []
        for tag in soup.findAll("td", class_="icoLineType4"):
            
            if tag.find(class_="pass-ico ico-color2"):
                seeing.append("25%")
            elif(tag.find(class_="pass-ico ico-color1") ):
                seeing.append("12,5%")
            elif tag.find(class_="pass-ico ico-color0") :
                seeing.append("0%")

        print(fName)
        wb = opx.Workbook()
        ws = wb.active
        try: 
            ws.title = soup.title.string
        except ValueError:
            print("ValueError")
            pass
        """ws['A'] = services
        ws['B'] = wheelChair
        ws['C'] = walkStick
        ws['D'] = hearing
        ws['E'] = seeing
        ws['F'] = summing"""
        
        ws.append((" ", "Коляска", "Трость", "Слух", "Зрение", "Сумма"))
        for row in zip(services, wheelChair, walkStick, hearing, seeing, summing):
            ws.append(row)
        chdir("{}/tables".format(homeDirectory))
        wb.save('{}.xlsx'.format(fName.replace(".html", "")))
        wb.close()
        chdir("{}/downloads".format(homeDirectory))

    print("{}%".format(i/len(onlyfiles)*100))
    